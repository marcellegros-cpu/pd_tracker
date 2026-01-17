"""
Export module for PD Tracker.

Generates reports in CSV, Excel, and PDF formats.
Supports exporting:
- Medication adherence data
- Symptom history
- Sleep logs
- Exercise logs
- Comprehensive reports (all data)
"""

import csv
import io
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from .database import get_connection
from .models import get_medication_status_today, get_all_medications
from .symptoms import get_symptoms_range
from .sleep import get_sleep_logs, get_sleep_stats
from .exercise import get_exercise_logs, get_exercise_stats


# Default export directory
EXPORT_DIR = Path(__file__).parent.parent / "exports"


def ensure_export_dir():
    """Create export directory if it doesn't exist."""
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    return EXPORT_DIR


# ============================================================
# DATA RETRIEVAL
# ============================================================

def get_medication_history(start_date: date, end_date: date) -> list:
    """Get medication dose history for date range."""
    conn = get_connection()
    cursor = conn.cursor()

    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date + timedelta(days=1), datetime.min.time())

    cursor.execute("""
        SELECT d.*, m.name as medication_name, m.dosage
        FROM doses_taken d
        JOIN medications m ON d.medication_id = m.id
        WHERE d.taken_time >= ? AND d.taken_time < ?
        ORDER BY d.taken_time DESC
    """, (start_dt, end_dt))

    rows = cursor.fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_all_data(start_date: date, end_date: date) -> dict:
    """Get all tracking data for a date range."""
    return {
        'medications': get_medication_history(start_date, end_date),
        'symptoms': [dict(s) for s in get_symptoms_range(start_date, end_date)],
        'sleep': [dict(s) for s in get_sleep_logs((end_date - start_date).days + 1)],
        'exercise': [dict(e) for e in get_exercise_logs((end_date - start_date).days + 1)],
    }


# ============================================================
# CSV EXPORT
# ============================================================

def export_csv(data_type: str, start_date: date, end_date: date,
               filename: str = None) -> Path:
    """
    Export data to CSV file.

    Args:
        data_type: 'medications', 'symptoms', 'sleep', 'exercise', or 'all'
        start_date: Start of date range
        end_date: End of date range
        filename: Optional custom filename

    Returns:
        Path to the created CSV file
    """
    ensure_export_dir()

    if filename is None:
        filename = f"pd_tracker_{data_type}_{start_date}_{end_date}.csv"

    filepath = EXPORT_DIR / filename
    all_data = get_all_data(start_date, end_date)

    if data_type == 'all':
        # For 'all', create multiple CSVs
        files = []
        for dtype in ['medications', 'symptoms', 'sleep', 'exercise']:
            if all_data[dtype]:
                f = export_csv(dtype, start_date, end_date)
                files.append(f)
        return files[0].parent if files else EXPORT_DIR

    data = all_data.get(data_type, [])

    if not data:
        # Create empty file with headers
        with open(filepath, 'w', newline='') as f:
            f.write(f"No {data_type} data for this period\n")
        return filepath

    # Write CSV
    with open(filepath, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)

    return filepath


# ============================================================
# EXCEL EXPORT
# ============================================================

def export_excel(start_date: date, end_date: date, filename: str = None) -> Path:
    """
    Export all data to Excel file with multiple sheets.

    Args:
        start_date: Start of date range
        end_date: End of date range
        filename: Optional custom filename

    Returns:
        Path to the created Excel file
    """
    ensure_export_dir()

    if filename is None:
        filename = f"pd_tracker_report_{start_date}_{end_date}.xlsx"

    filepath = EXPORT_DIR / filename
    all_data = get_all_data(start_date, end_date)

    # Create workbook
    wb = Workbook()

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def style_sheet(ws, df):
        """Apply styling to a worksheet."""
        # Header styling
        for col_num, cell in enumerate(ws[1], 1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Data styling
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # Remove default sheet
    wb.remove(wb.active)

    # Medications sheet
    if all_data['medications']:
        ws = wb.create_sheet("Medications")
        df = pd.DataFrame(all_data['medications'])
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        style_sheet(ws, df)

    # Symptoms sheet
    if all_data['symptoms']:
        ws = wb.create_sheet("Symptoms")
        df = pd.DataFrame(all_data['symptoms'])
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        style_sheet(ws, df)

    # Sleep sheet
    if all_data['sleep']:
        ws = wb.create_sheet("Sleep")
        df = pd.DataFrame(all_data['sleep'])
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        style_sheet(ws, df)

    # Exercise sheet
    if all_data['exercise']:
        ws = wb.create_sheet("Exercise")
        df = pd.DataFrame(all_data['exercise'])
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        style_sheet(ws, df)

    # Summary sheet
    ws = wb.create_sheet("Summary", 0)
    ws.append(["PD Tracker Report"])
    ws.append([f"Period: {start_date} to {end_date}"])
    ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append([])
    ws.append(["Data Summary"])
    ws.append(["Medication doses logged:", len(all_data['medications'])])
    ws.append(["Symptom entries:", len(all_data['symptoms'])])
    ws.append(["Sleep logs:", len(all_data['sleep'])])
    ws.append(["Exercise sessions:", len(all_data['exercise'])])

    # Style summary
    ws['A1'].font = Font(bold=True, size=16)
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

    wb.save(filepath)
    return filepath


# ============================================================
# PDF EXPORT
# ============================================================

def export_pdf(start_date: date, end_date: date, filename: str = None) -> Path:
    """
    Export summary report to PDF.

    Args:
        start_date: Start of date range
        end_date: End of date range
        filename: Optional custom filename

    Returns:
        Path to the created PDF file
    """
    ensure_export_dir()

    if filename is None:
        filename = f"pd_tracker_report_{start_date}_{end_date}.pdf"

    filepath = EXPORT_DIR / filename
    all_data = get_all_data(start_date, end_date)

    # Create PDF
    doc = SimpleDocTemplate(str(filepath), pagesize=letter,
                           rightMargin=72, leftMargin=72,
                           topMargin=72, bottomMargin=72)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
    )
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        spaceBefore=20,
        spaceAfter=10,
    )

    elements = []

    # Title
    elements.append(Paragraph("PD Tracker Report", title_style))
    elements.append(Paragraph(f"Period: {start_date} to {end_date}", styles['Normal']))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))

    # Summary section
    elements.append(Paragraph("Summary", heading_style))
    summary_data = [
        ["Metric", "Value"],
        ["Medication doses logged", str(len(all_data['medications']))],
        ["Symptom entries", str(len(all_data['symptoms']))],
        ["Sleep logs", str(len(all_data['sleep']))],
        ["Exercise sessions", str(len(all_data['exercise']))],
    ]
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    # Medications section
    if all_data['medications']:
        elements.append(Paragraph("Medication Log", heading_style))

        # Group by medication
        med_counts = {}
        for m in all_data['medications']:
            name = m['medication_name']
            if name not in med_counts:
                med_counts[name] = 0
            med_counts[name] += 1

        med_data = [["Medication", "Doses"]]
        for name, count in sorted(med_counts.items()):
            med_data.append([name, str(count)])

        med_table = Table(med_data, colWidths=[3*inch, 1.5*inch])
        med_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]))
        elements.append(med_table)
        elements.append(Spacer(1, 20))

    # Symptoms section
    if all_data['symptoms']:
        elements.append(Paragraph("Symptom Summary", heading_style))

        on_count = sum(1 for s in all_data['symptoms'] if s.get('on_off_state') == 'on')
        off_count = sum(1 for s in all_data['symptoms'] if s.get('on_off_state') == 'off')
        trans_count = sum(1 for s in all_data['symptoms'] if s.get('on_off_state') == 'transitioning')

        symptom_data = [
            ["State", "Count"],
            ["ON", str(on_count)],
            ["OFF", str(off_count)],
            ["Transitioning", str(trans_count)],
        ]
        symptom_table = Table(symptom_data, colWidths=[2*inch, 1.5*inch])
        symptom_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(symptom_table)
        elements.append(Spacer(1, 20))

    # Sleep section
    if all_data['sleep']:
        elements.append(Paragraph("Sleep Summary", heading_style))

        total_nights = len([s for s in all_data['sleep'] if s.get('wake_time')])
        qualities = [s['quality'] for s in all_data['sleep'] if s.get('quality')]
        avg_quality = sum(qualities) / len(qualities) if qualities else 0

        sleep_data = [
            ["Metric", "Value"],
            ["Nights logged", str(total_nights)],
            ["Avg quality", f"{avg_quality:.1f}/10" if avg_quality else "N/A"],
        ]
        sleep_table = Table(sleep_data, colWidths=[2*inch, 1.5*inch])
        sleep_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(sleep_table)
        elements.append(Spacer(1, 20))

    # Exercise section
    if all_data['exercise']:
        elements.append(Paragraph("Exercise Summary", heading_style))

        total_minutes = sum(e['duration_minutes'] for e in all_data['exercise'])
        total_sessions = len(all_data['exercise'])

        exercise_data = [
            ["Metric", "Value"],
            ["Total sessions", str(total_sessions)],
            ["Total minutes", str(total_minutes)],
        ]
        exercise_table = Table(exercise_data, colWidths=[2*inch, 1.5*inch])
        exercise_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(exercise_table)

    # Build PDF
    doc.build(elements)
    return filepath


# ============================================================
# CONVENIENCE FUNCTIONS
# ============================================================

def export_last_week(format: str = 'pdf') -> Path:
    """Export data from the last 7 days."""
    end_date = date.today()
    start_date = end_date - timedelta(days=6)

    if format == 'csv':
        return export_csv('all', start_date, end_date)
    elif format == 'excel':
        return export_excel(start_date, end_date)
    else:
        return export_pdf(start_date, end_date)


def export_last_month(format: str = 'pdf') -> Path:
    """Export data from the last 30 days."""
    end_date = date.today()
    start_date = end_date - timedelta(days=29)

    if format == 'csv':
        return export_csv('all', start_date, end_date)
    elif format == 'excel':
        return export_excel(start_date, end_date)
    else:
        return export_pdf(start_date, end_date)
